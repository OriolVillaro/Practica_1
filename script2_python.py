#! /usr/vin/env python3

import openpyxl
import sys
import os
import fnmatch

repetits=0
#desti_dir=input("En quin directori són els fitxers a analitzar?\n") #Demanem el directori on són els fitxers i el canviem 
desti_dir= "/home/ori/FSO/Pràctica_1" #FICAR L'INPUT A LA VERSIÓ FINAL
os.chdir(desti_dir)	#Input estàndard: /home/ori/FSO/Pràctica_1

files = []
num_files=0

dicc_total = {}

listOfFiles = os.listdir('.')

pattern = "*.xlsx"	#Només agafem els fitxers amb extensió d'Excel
for entry in listOfFiles:
	if fnmatch.fnmatch(entry, pattern):
		files.append(entry)
		
for num_files in range(len(files)):
	print("\nFitxer:"+files[num_files]+"\n")
	
	f=open("llista_noms_"+files[num_files]+".txt","w+")
	f.write("NOM REAL:\t\t NOM ANÒNIM:\n\n")

	wb = openpyxl.load_workbook(files[num_files])
	num_sheets = len(wb.sheetnames)-4 #Serà el nombre de 'Sheet ?', les totals menys 4 a cada fitxer
	tabs=wb.active
	tabs = wb['Overview']
	#sheets=tabs[0] #Agafem la pàgina de Overview
	num_jug = tabs["B4"].value
	num, jug = num_jug.split() #Aconseguim el número de jugadors pel bucle.
	int_num = int(num)

	num_preg = tabs["B5"].value[0:2]
	int_preg = int(num_preg)

	i,j,k = 4,1,1 #Inicialitzem la primera fila on hi han noms i índexs
	int_num_real = int_num + i #Ens assegurem que arribi al final de la llista de noms
	diccionari = {}

#CREACIÓ DEL FITXER DE NOMS ANONIMITZATS I DEL DICCIONARI:

	while(i<int_num_real): #Bucle per les dues primeres sheets. Separaem en 3 bucles degut a que els noms comencen en llocs diferents
		
		tabs = wb['Final Scores']
		nom = tabs.cell(row=i, column=2).value
		nom_nou = hash(nom)		#Anonimitzem els noms mitjançant el mètode hash
		f.write(str(nom) +"\t\t" + str(nom_nou) + "\n") #Escrivim al fitxer els noms originals i anonimitzats
		diccionari[nom]=nom_nou
		
		if nom in dicc_total:	#Si el jugador està present a un altra fitxer no el tornarem a afegir
			repetits+=1
			print(nom)
		else:
			dicc_total[nom]=nom_nou	

		i+=1
		j+=1
	
	i,j = 4,1
	print("\t-Diccionari creat")

#BUCLES DE CANVI DE NOM:

	while(i<int_num_real): #Bucle per les dues primeres sheets. Separaem en 3 bucles degut a que els noms comencen en llocs diferents
		
		nom_nou = diccionari[tabs.cell(row=i, column=2).value]
		tabs = wb['Final Scores']
		tabs.cell(row=i, column=2).value=nom_nou #Canviem el nom al fitxer
		tabs = wb['Question Summary'] #Canviem de tab per modifcar les dues
		tabs.cell(row=i, column=2).value=nom_nou 
		i+=1
		j+=1
		
	f.close()
	i=15
	int_num_real = int_num + i

	while(k<=num_sheets):

		i,j = 15,1
		tabs = wb['Question %d' % k]
		
		while(i<(int_num_real)): #Segon bucle. En un futur fer un split a tabs[-2] per saber el número de Question
			nom_nou = diccionari[tabs.cell(row=i, column=1).value]
			tabs.cell(row=i, column=1).value=nom_nou #Canviem el nom al fitxer
			i+=1
			j+=1
		k+=1

	i,j = 2,1
	int_num_real = int_num + i
	tabs = wb['RawReportData Data']


	while(i<(int_num*int_preg+2)): #Tercer bucle
		
		if(j>int_num): j=1
		nom_nou = diccionari[tabs.cell(row=i, column=9).value]
		tabs.cell(row=i, column=9).value=nom_nou #Canviem el nom al fitxer
		i+=1
		j+=1
	
	files[num_files]="Nou_"+files[num_files]
	wb.save(files[num_files])
	print("\t-Noms anonimitzats")
	
#COMPROVACIÓ QUE ELS PUNTS, LES PREGUNTES CORRECTES I LES INCORRECTES ESTAN BÉ

	wb = openpyxl.load_workbook(files[num_files])

	i,k,l,m=4,1,0,4
	int_num_real = int_num + i
	punts_total, correctes, incorrectes=0,0,0	#Inicialització de variables i índexs que utilitzarem

	f_modif=open("modificacions_"+files[num_files]+".txt", "w+")
	f_modif.write("NOM:\t\t\tPUNTS ORIGINALS:\tPUNTS CALCULATS:\n\n")	#Capceleres dels fitxers de modificació
	
	num_modif=0
	
	while(k<=int_num): #k és el número de jugadors
		
		tabs = wb['Final Scores']
		punts_originals = tabs.cell(row=i, column=3).value	#Obtenim els punts que el fitxer diu que té el jugador, així com el número de respostes correctes i incorrectes
		correctes_originals = tabs.cell(row=i, column=4).value
		incorrectes_originals = tabs.cell(row=i, column=5).value
		#print(punts_originals,correctes_originals,incorrectes_originals)
		tabs = wb['Question Summary']
			
		while(l<int(num_preg)): #Per cada pregunta aconseguirem els punts obtinguts i els sumarem a un acumulador que acabarà sent els punts totals
			
			#print(l, num_preg)
			punts_pregunta = tabs.cell(row=i, column=m).value
			#print(punts_pregunta)
			m=m+2
			punts_total=punts_total+punts_pregunta
			
			if(punts_pregunta>0): correctes=correctes+1		#Si els punts de la pregunta són > 0 serà correcte, si son 0 serà incorrecte
			else: incorrectes=incorrectes+1
		
			l+=1
		
		if(punts_originals != punts_total):		#Si hi ha alguna discrepància entre els punts originals i l'acumulador s'apuntarà al fitxer modificacions i es sobreescriurà el valor
			nom=tabs.cell(row=i, column=2).value
			#print(nom+"\t\t"+punts_originals+"\t\t"+punts_total+"\n")
			f_modif.write(str(nom) +"\t\t"+ str(punts_originals) +"\t\t\t"+ str(punts_total)+"\n")
			
			tabs = wb['Final Scores']
			tabs.cell(row=i, column=3).value = str(punts_total)
			num_modif+=1
		
		l,m=0,4
		i+=1
		k+=1
		
		correctes, incorrectes=0,0
		punts_total=0
	print("\t-Punts totals calculats, s'han realitzat %d modificacions\n" % num_modif)
	print("Procés complert\n\n")
	f_modif.close()
	wb.save(files[num_files])
